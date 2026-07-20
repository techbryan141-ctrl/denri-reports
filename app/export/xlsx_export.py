"""Renders the same report_data dict served by /api/report into a .xlsx
workbook - one sheet per major section, reusing the dashboard's already-
computed, already-formatted values verbatim (no metric is recalculated here).
"""
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(bold=True, size=13, color="1F3864")
SUBTITLE_FONT = Font(bold=True, size=11, color="1F3864")
BODY_FONT = Font(size=10)
MUTED_FONT = Font(size=10, italic=True, color="898781")
THIN_BORDER = Border(*(Side(style="thin", color="D9D9D9"),) * 4)

_DIR_COLOR = {"up": "0CA30C", "down": "D03B3B", "neutral": "52514E"}
_TONE_COLOR = {"good": "0CA30C", "warning": "B88600", "critical": "D03B3B", "neutral": "52514E"}
PERIOD_TITLES = {"day": "Daily", "week": "Weekly", "month": "Monthly"}
_PRIORITY_COLOR = {"CRITICAL": "D03B3B", "HIGH": "B88600", "MEDIUM": "52514E", "WIN": "0CA30C"}


def _title(ws, row, text):
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = TITLE_FONT
    return row + 2


def _subtitle(ws, row, text):
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = SUBTITLE_FONT
    return row + 1


def _note(ws, row, text):
    if not text:
        return row
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = MUTED_FONT
    return row + 2


def _write_table(ws, row, headers, rows, getters, color_specs=None):
    if not rows:
        return _note(ws, row, "No data for this period.")

    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="left")
    row += 1

    for r in rows:
        is_total = str(r.get("shop", r.get("metric", ""))).strip().upper() == "TOTAL"
        for c, getter in enumerate(getters, start=1):
            cell = ws.cell(row=row, column=c, value=str(getter(r)))
            cell.border = THIN_BORDER
            cell.font = Font(size=10, bold=is_total)
            if color_specs and c - 1 in color_specs:
                color = color_specs[c - 1](r)
                if color:
                    cell.font = Font(size=10, bold=is_total, color=color)
        row += 1

    return row + 1


def _dir_color(row):
    return _DIR_COLOR.get(row.get("dir"))


def _tone_color(row):
    return _TONE_COLOR.get(row.get("status_tone"))


def _autosize(ws, max_col=10):
    for col_idx in range(1, max_col + 1):
        letter = get_column_letter(col_idx)
        longest = 0
        for cell in ws[letter]:
            if cell.value:
                longest = max(longest, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max(longest + 2, 10), 45)


def _summary_sheet(wb, report):
    ws = wb.active
    ws.title = "Summary"
    meta = report["meta"]

    row = _title(ws, 1, f"{PERIOD_TITLES.get(meta['period_type'], '')} Sales Performance Report")
    ws.cell(row=row, column=1, value=f"{meta['period_label']} | {meta['date_range']} ({meta['days']} day(s))").font = BODY_FONT
    row += 1
    ws.cell(row=row, column=1, value=f"Compared to: {meta['compared_to']}").font = MUTED_FONT
    row += 1
    ws.cell(row=row, column=1, value=f"Generated: {meta['generated_on']}").font = MUTED_FONT
    row += 2

    row = _subtitle(ws, row, "Key Performance Indicators")
    row = _write_table(
        ws, row, ["Label", "Value", "Change", "Tone"], report["kpi_strip"],
        [lambda r: r["label"], lambda r: r["value"], lambda r: r["sub"], lambda r: r["tone"]],
        color_specs={3: lambda r: _DIR_COLOR.get(r["tone"])},
    )

    ws.cell(row=row, column=1, value=report["context_note"]).font = MUTED_FONT
    row += 2
    ws.cell(row=row, column=1, value=report["exec_summary"]["narrative"]).font = BODY_FONT
    ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)

    _autosize(ws, max_col=4)


def _exec_summary_sheet(wb, data):
    ws = wb.create_sheet("Executive Summary")
    row = _title(ws, 1, "1. Executive Summary")
    row = _write_table(
        ws, row, ["Metric", "Current", "Previous", "Change"], data["metrics"],
        [lambda r: r["metric"], lambda r: r["current"], lambda r: r["previous"], lambda r: r["change"]],
        color_specs={3: _dir_color},
    )
    row = _subtitle(ws, row, "Flags")
    for flag in data["flags"]:
        ws.cell(row=row, column=1, value=f"[{flag['kind'].upper()}] {flag['text']}").font = BODY_FONT
        row += 1
    row += 1
    row = _note(ws, row, data["meeting_note"])
    _autosize(ws)


def _customer_metrics_sheet(wb, data):
    ws = wb.create_sheet("Customer Metrics")
    row = _title(ws, 1, "2. Customer Metrics")
    row = _write_table(
        ws, row, ["Metric", "Current", "Previous", "Change"], data["comparison"],
        [lambda r: r["metric"], lambda r: r["current"], lambda r: r["previous"], lambda r: r["change"]],
        color_specs={3: _dir_color},
    )
    row = _subtitle(ws, row, "2.1 Purchase Frequency Distribution")
    row = _write_table(
        ws, row, ["Purchase Frequency", "Customer Count", "Percentage", "Read"], data["frequency"],
        [lambda r: r["band"], lambda r: r["count"], lambda r: r["pct"], lambda r: r["read"]],
    )
    row = _subtitle(ws, row, "2.2 Channel Overview")
    row = _write_table(
        ws, row, ["Channel", "Orders", "Share", "Avg Spend", "Note"], data["channels"],
        [lambda r: r["channel"], lambda r: r["orders"], lambda r: r["share"], lambda r: r["avg_spend"], lambda r: r["note"]],
    )
    _note(ws, row, data["meeting_note"])
    _autosize(ws)


def _store_performance_sheet(wb, data):
    ws = wb.create_sheet("Store Performance")
    row = _title(ws, 1, "3. Store Performance")
    row = _write_table(
        ws, row, ["Shop", "Current", "Previous", "Change", "Change %", "New", "Repeat"], data["rows"],
        [lambda r: r["shop"], lambda r: r["current"], lambda r: r["previous"], lambda r: r["change"],
         lambda r: r["change_pct"], lambda r: r["new"], lambda r: r["repeat"]],
        color_specs={3: _dir_color, 4: _dir_color},
    )
    row = _subtitle(ws, row, "3.1 Channel Mix by Location")
    row = _write_table(
        ws, row, ["Shop", "Walk-in", "Online", "Activation", "Total", "Online %"], data["channel_mix"],
        [lambda r: r["shop"], lambda r: r["walkin"], lambda r: r["online"], lambda r: r["activation"],
         lambda r: r["total"], lambda r: r["online_pct"]],
    )
    _note(ws, row, data["meeting_note"])
    _autosize(ws)


def _gender_performance_sheet(wb, data):
    ws = wb.create_sheet("Gender Performance")
    row = _title(ws, 1, "4. Gender Performance Analysis")
    if data.get("ratio"):
        ws.cell(row=row, column=1, value=f"Ratio: {data['ratio']['ratio_text']}").font = Font(bold=True, size=10, color="2A78D6")
        row += 2

    row = _subtitle(ws, row, "4.1 Overall Gender Split")
    row = _write_table(
        ws, row, ["Gender", "Count", "%", "Change"], data["overall"],
        [lambda r: r["gender"], lambda r: r["count"], lambda r: r["pct"], lambda r: r["change"]],
        color_specs={3: _dir_color},
    )

    row = _subtitle(ws, row, "4.2 Gender by Location")
    cols = data["by_location"]["columns"]
    headers = ["Shop"] + cols + ["Total"]
    getters = [lambda r: r["shop"]] + [(lambda r, c=c: r[c.lower()]) for c in cols] + [lambda r: r["total"]]
    row = _write_table(ws, row, headers, data["by_location"]["rows"], getters)

    _note(ws, row, data["meeting_note"])
    _autosize(ws)


def _traffic_sheet(wb, data):
    ws = wb.create_sheet("Traffic")
    row = _title(ws, 1, "5. Foot & Online Traffic Analysis")
    if data.get("data_gap_note"):
        row = _note(ws, row, f"DATA GAP: {data['data_gap_note']}")
    row = _write_table(
        ws, row,
        ["Shop", "Walk-in Purchased", "Walk-in Total", "Conv. Rate", "Online", "Activation", "Total Customers"],
        data["rows"],
        [lambda r: r["shop"], lambda r: r["walkin_purchased"], lambda r: r["walkin_total"], lambda r: r["conv_rate"],
         lambda r: r["online"], lambda r: r["activation"], lambda r: r["total_customers"]],
    )
    _note(ws, row, data["meeting_note"])
    _autosize(ws)


def _revenue_sheet(wb, data):
    ws = wb.create_sheet("Revenue Analysis")
    row = _title(ws, 1, "6. Revenue Analysis")
    row = _write_table(
        ws, row, ["Label", "Value", "Change"], data["kpi_cards"],
        [lambda r: r["label"], lambda r: r["value"], lambda r: r["sub"]],
        color_specs={2: lambda r: _DIR_COLOR.get(r["tone"])},
    )
    row = _write_table(
        ws, row, ["Metric", "Current", "Previous", "Change"], data["comparison"],
        [lambda r: r["metric"], lambda r: r["current"], lambda r: r["previous"], lambda r: r["change"]],
        color_specs={3: _dir_color},
    )
    row = _subtitle(ws, row, "6.1 Revenue Bridge")
    row = _write_table(
        ws, row, ["Period", "Revenue", "Customers", "Avg Spend", "Trend"], data["bridge"],
        [lambda r: r["period"], lambda r: r["revenue"], lambda r: r["customers"], lambda r: r["avg_spend"], lambda r: r["trend"]],
        color_specs={4: _dir_color},
    )
    _note(ws, row, data["meeting_note"])
    _autosize(ws)


def _data_quality_sheet(wb, data):
    ws = wb.create_sheet("Data Quality")
    row = _title(ws, 1, "7. Data Quality (KYC)")
    row = _write_table(
        ws, row, ["Metric", "Current", "Current %", "Previous", "Previous %", "Change"], data["comparison"],
        [lambda r: r["metric"], lambda r: r["current"], lambda r: r["current_pct"], lambda r: r["previous"],
         lambda r: r["previous_pct"], lambda r: r["change"]],
        color_specs={5: _dir_color},
    )
    if data.get("store_number_by_shop"):
        row = _subtitle(ws, row, "7.1 Store Numbers Used by Location")
        row = _write_table(
            ws, row, ["Shop", "Unique Numbers", "Occurrences"], data["store_number_by_shop"],
            [lambda r: r["shop"], lambda r: r["unique_numbers"], lambda r: r["occurrences"]],
        )
    row = _note(ws, row, data["score_note"])
    _note(ws, row, data["meeting_note"])
    _autosize(ws)


def _feedback_sheet(wb, data):
    ws = wb.create_sheet("Feedback")
    row = _title(ws, 1, "8. Customer Feedback")
    row = _write_table(
        ws, row, ["Metric", "Current", "Previous", "Change %", "Status"], data["comparison"],
        [lambda r: r["metric"], lambda r: r["current"], lambda r: r["previous"], lambda r: r["change"], lambda r: r["status"]],
        color_specs={3: _dir_color, 4: _tone_color},
    )
    if data.get("target"):
        row = _note(ws, row, f"TARGET: {data['target']['note']}")

    row = _subtitle(ws, row, "8.1 Store Breakdown")
    row = _write_table(
        ws, row, ["Shop", "Current", "Previous", "Change"], data["store_breakdown"],
        [lambda r: r["shop"], lambda r: r["current"], lambda r: r["previous"], lambda r: r["change"]],
        color_specs={3: _dir_color},
    )

    if data.get("feedback_links"):
        row = _subtitle(ws, row, "8.2 Feedback Link Targets by Location")
        row = _write_table(
            ws, row, ["Shop", "Links Sent", "Online", "Achv %", "Gap", "Status", "Walk-Ins"], data["feedback_links"],
            [lambda r: r["shop"], lambda r: r["links_sent"], lambda r: r["online"], lambda r: r["achv_pct"],
             lambda r: r["gap"], lambda r: r["status"], lambda r: r["walkins"]],
            color_specs={5: _tone_color},
        )

    _note(ws, row, data["meeting_note"])
    _autosize(ws)


def _priorities_sheet(wb, data):
    ws = wb.create_sheet("Summary & Key Notes")
    row = _title(ws, 1, "9. Summary & Key Notes")
    dept_labels = {"marketing": "Marketing", "sales": "Sales", "production": "Production", "data": "Data"}

    for key, label in dept_labels.items():
        dept = data[key]
        row = _subtitle(ws, row, label)
        ws.cell(row=row, column=1, value=dept["summary"]).font = MUTED_FONT
        row += 1
        if dept.get("unavailable"):
            ws.cell(row=row, column=1, value=dept["note"]).font = MUTED_FONT
            row += 2
            continue
        for item in dept["items"]:
            cell = ws.cell(row=row, column=1, value=f"[{item['priority']}] {item['text']}")
            cell.font = Font(size=10, color=_PRIORITY_COLOR.get(item["priority"], "000000"))
            row += 1
        row += 1

    _autosize(ws)


def build_xlsx(report: dict) -> io.BytesIO:
    wb = Workbook()
    _summary_sheet(wb, report)
    _exec_summary_sheet(wb, report["exec_summary"])
    _customer_metrics_sheet(wb, report["customer_metrics"])
    _store_performance_sheet(wb, report["store_performance"])
    _gender_performance_sheet(wb, report["gender_performance"])
    _traffic_sheet(wb, report["traffic"])
    _revenue_sheet(wb, report["revenue_analysis"])
    _data_quality_sheet(wb, report["data_quality"])
    _feedback_sheet(wb, report["feedback"])
    _priorities_sheet(wb, report["priorities"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
